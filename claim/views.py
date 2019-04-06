from django.shortcuts import render
from matplotlib.pyplot import *
from datetime import datetime
import PyPDF2
from openpyxl import Workbook
from io import StringIO
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage

import ibm_db

invoice_ids = set()

posts = [
    {'a': 'Welcome to Health Insurance Claim Portal',
     'b': '1. Secure & Faster Claim Process with Signature verification  ',
     'c': '2. Invoice Validation Against The Claim',
     'd': '3. Login with Provided Credentials And Upload Documents For a New Claim',
     'e': 'Login and Create a New Claim?',
     'f': 'File a New Claim?'
     }
]

def home(request):
    context = {
        'posts': posts
    }
    return render(request, 'claim/home.html', context)


def about(request):
    return render(request, 'claim/about.html', {'title': 'About'})


def login(request):
    from subprocess import call
    call(["python", "new.py"])
    return login


def account(request):
    if request.method == 'POST':
        claim_number = int(request.POST['policy'])
        claim_amount = int(request.POST['amount'])

        prescription = request.FILES['document']

        invoice = request.FILES['invoice']
        signature = request.POST['signature']

        # fs = FileSystemStorage()
        # file_name = fs.save(uploaded_file.name, uploaded_file)
        # uploaded_file_url = fs.url(file_name)

        if not (verifyClaimAgainstDBDetails(request.user.username, claim_number, claim_amount, signature)):
            return render(request, 'claim/result.html', {'title': 'Failure'})

        if not (verifyPrescription(prescription,request.user.username)):
            return render(request, 'claim/failDocument.html', {'title': 'Failure'})

        if not (verifyInvoice(claim_amount, invoice)):
            return render(request, 'claim/failInvoice.html', {'title': 'Failure'})

        if (isFraud(request.user.username,claim_number,claim_amount)):
            return render(request, 'claim/fraudInvoice.html', {'title': 'Failure'})

        else:
            updateTransaction(request.user.username, claim_amount)
            return render(request, 'claim/success.html', {'title': 'account'})


    else:
        return render(request, 'claim/account.html', {'title': 'account'})


def isFraud(username,pnumber,amount):

    mydb = ibm_db.connect('DATABASE=BLUDB;'
                          'HOSTNAME=dashdb-txn-sbox-yp-lon02-01.services.eu-gb.bluemix.net;'  # 127.0.0.1 or localhost works if it's local
                          'PORT=50000;'
                          'PROTOCOL=TCPIP;'
                          'UID=trp11982;'
                          'PWD=q144^vv5vb7pdv71;', '', '')

    for invoice_i in invoice_ids:
        sql = "SELECT * FROM invoices where invoice_id = " + "'" + invoice_i + "'"
        print(sql)

        stmt = results(ibm_db.exec_immediate(mydb, sql))

        if (stmt.__len__() > 0):
            d = str(datetime.today().strftime('%Y-%m-%d'))
            sql2 = "insert into Frauds values(" + "'" + username + "'," +str(pnumber)+","+str(amount)+",'"+d+ "')"
            ibm_db.exec_immediate(mydb, sql2)
            return True

    return False

    # converts pdf, returns its text content as a string
def convert(infile, pages=None):
    if not pages:
        pagenums = set()
    else:
        pagenums = set(pages)

    output = StringIO()
    manager = PDFResourceManager()
    converter = TextConverter(manager, output, laparams=LAParams())
    interpreter = PDFPageInterpreter(manager, converter)

    for page in PDFPage.get_pages(infile, pagenums):
        interpreter.process_page(page)
    infile.close()
    converter.close()
    text = output.getvalue()
    output.close
    return text

def verifyPrescription(uploaded_file,username):

    txt = convert(uploaded_file)
    # open a text document magic.txt in downloads and write all the text from pdf
    text = str(txt)
    text = text.lower()
    user = username.lower()

    if user not in text:
        return False

    if ('dr.' in text) or ('dr' in text) or ('md' in text) or ('m.d.' in text) or 'medical' in text or 'hospital' in text:
        return True

    return False

def results(command):
    ret = []
    result = ibm_db.fetch_assoc(command)
    while result:
        # This builds a list in memory. Theoretically, if there's a lot of rows,
        # we could run out of memory. In practice, I've never had that happen.
        # If it's ever a problem, you could use
        #     yield result
        # Then this function would become a generator. You lose the ability to access
        # results by index or slice them or whatever, but you retain
        # the ability to iterate on them.
        ret.append(result)
        result = ibm_db.fetch_assoc(command)
    return ret


def verifyClaimAgainstDBDetails(username, claim_number, claim_amount, uploaded_file2):
    mydb = ibm_db.connect('DATABASE=BLUDB;'
                          'HOSTNAME=dashdb-txn-sbox-yp-lon02-01.services.eu-gb.bluemix.net;'  # 127.0.0.1 or localhost works if it's local
                          'PORT=50000;'
                          'PROTOCOL=TCPIP;'
                          'UID=trp11982;'
                          'PWD=q144^vv5vb7pdv71;', '', '')

    sql = "SELECT * FROM UserDetails where username = " + "'" + username + "'"
    print(sql)

    stmt = results(ibm_db.exec_immediate(mydb, sql))

    if (stmt.__len__() < 1):
        return False

    for x in stmt:
        print(x)

    data = str(stmt).split(' ')

    data[3] = data[3].replace(',', '')
    data[7] = data[7].replace(',', '')
    if int(data[3]) != claim_number:
        return False;

    if int(data[7]) < claim_amount:
        return False;

    return (validateSignature(uploaded_file2, data[9]))
    return True

def validateSignature(signature_uploaded_by_user, URL):
    try:
        from PIL import Image
        import requests
        from io import BytesIO

        URL = URL[1:-1]
        URL = URL.replace('}', '')
        i1 = Image.open(requests.get(URL, stream=True).raw)

        URL = signature_uploaded_by_user
        i2 = Image.open(requests.get(URL, stream=True).raw)

        pairs = zip(i1.getdata(), i2.getdata())
        if len(i1.getbands()) == 1:
            # for gray-scale jpegs
            dif = sum(abs(p1 - p2) for p1, p2 in pairs)
        else:
            dif = sum(abs(c1 - c2) for p1, p2 in pairs for c1, c2 in zip(p1, p2))
        ncomponents = i1.size[0] * i1.size[1] * 3
        n = (dif / 255.0 * 100) / ncomponents
        if n <= 4.3:
            return True
        else:
            return False
    except:
        return False


def pdf2Invoie(request, uploaded_file):
    input_file = uploaded_file
    output_file = 'C:\\Users\\KHALID KHAN\\Downloads\\invoice.xlsx'

    pdf_file = open(input_file, 'rb')
    input_pdf = PyPDF2.PdfFileReader(pdf_file)

    main_list = ['From',
                 'To',
                 'Invoice Number',
                 'Order Number',
                 'Invoice Date',
                 'Due Date',
                 'Total Due',
                 'Quantity',
                 'Service',
                 'Rate',
                 'Adjust',
                 'Sub Total',
                 '!"#$%&#']

    wb = Workbook()
    ws = wb.active

    row_num = 1
    column_num = 1
    for i in range(len(main_list) - 1):
        field = main_list[i]
        ws.cell(row=row_num, column=column_num, value=field)
        column_num += 1

    total_pages = input_pdf.getNumPages()

    row_num = 2
    for i in range(total_pages):
        page = input_pdf.getPage(i)
        page_content = page.extractText()
        column_num = 1
        for i in range(len(main_list) - 1):
            field = main_list[i]
            next_field = main_list[i + 1]

            field_pos = page_content.find(field)
            next_field_pos = page_content.find(next_field)

            field_value_start_pos = field_pos + len(field)
            field_value_end_pos = next_field_pos

            field_value = page_content[field_value_start_pos:field_value_end_pos]
            ws.cell(row=row_num, column=column_num, value=field_value)
            column_num += 1
        row_num += 1

    pdf_file.close()

    wb.save(output_file)


def result(request):
    return render(request, 'claim/result.html', {'title': 'Failed'})


def invoiceReq(request):
    return render(request, 'claim/failInvoice.html.html', {'title': 'Failed'})


def document(request):
    return render(request, 'claim/failDocument.html', {'title': 'Failed'})


def success(request):
    return render(request, 'claim/success.html', {'title': 'Success'})


def verifyInvoice(claim_amount, pdf_file):
    try:
        input_pdf = PyPDF2.PdfFileReader(pdf_file)

        main_list = ['From',
                     'To',
                     'Invoice Number',
                     'Order Number',
                     'Invoice Date',
                     'Due Date',
                     'Total Due',
                     'Quantity',
                     'Service',
                     'Rate',
                     'Adjust',
                     'Sub Total',
                     '!"#$%&#']

        wb = Workbook()
        ws = wb.active

        row_num = 1
        column_num = 1
        for i in range(len(main_list) - 1):
            field = main_list[i]
            ws.cell(row=row_num, column=column_num, value=field)
            column_num += 1

        total_pages = input_pdf.getNumPages()

        row_num = 2
        for i in range(total_pages):
            page = input_pdf.getPage(i)
            page_content = page.extractText()
            column_num = 1
            for i in range(len(main_list) - 1):
                field = main_list[i]
                next_field = main_list[i + 1]

                field_pos = page_content.find(field)
                next_field_pos = page_content.find(next_field)

                field_value_start_pos = field_pos + len(field)
                field_value_end_pos = next_field_pos

                field_value = page_content[field_value_start_pos:field_value_end_pos]
                ws.cell(row=row_num, column=column_num, value=field_value)
                column_num += 1
            row_num += 1

        pdf_file.close()

        total = 0
        i = 0
        for row_cells in ws.iter_rows():
            if (i > 0):
                j = 0
                for cell in row_cells:
                    if (j == 2):
                        inv = cell.value
                        inv = inv.replace('\n', '')
                        inv = inv.replace(' ', '')
                        invoice_ids.add(inv)
                    if (j == 11):
                        val = cell.value
                        val = val.replace('\n', '')
                        val = val.replace(' ', '')
                        val = val.replace('$', '')
                        total = total + float(val)
                    j = j + 1
                    # print(cell.value)

            i = i + 1

        if (total == claim_amount):
            return True

        return False
    except:
        return False


def updateTransaction(username, claim_amount):

    mydb = ibm_db.connect('DATABASE=BLUDB;'
                         'HOSTNAME=dashdb-txn-sbox-yp-lon02-01.services.eu-gb.bluemix.net;'  # 127.0.0.1 or localhost works if it's local
                         'PORT=50000;'
                         'PROTOCOL=TCPIP;'
                         'UID=trp11982;'
                         'PWD=q144^vv5vb7pdv71;', '', '')

    sql = "SELECT * FROM UserDetails where username = " +  "'" + username + "'"
    print(sql)

    stmt = results(ibm_db.exec_immediate(mydb, sql))

    if(stmt.__len__()<1) :
        return False

    for x in stmt:
        print(x)

    data = str(stmt).split(' ')

    data[3] = data[3].replace(',','')
    data[5] = data[5].replace(',','')
    data[7] = data[7].replace(',', '')
    newClaimedAmount = claim_amount + int(data[5])

    sql = "update UserDetails set pAmountClaimed =" + str(newClaimedAmount) +" where username = " + "'" + username + "'"
    ibm_db.exec_immediate(mydb, sql)

    invoice_list = "("
    for inv in invoice_ids:
        sql2 = "insert into Invoices values("+"'"+inv+"'"+")"
        ibm_db.exec_immediate(mydb, sql2)


    return
